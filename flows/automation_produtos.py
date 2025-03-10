import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import schedule
import time

def weekly_data_processing():
    try:
        # pandas
        df = pd.read_excel("./dados/produtos(1).xlsx")
        
        required_columns = ["Preço Base Original", "Multiplicador Imposto", "Produtos"]
        for column in required_columns:
            if column not in df.columns:
                raise ValueError(f"A coluna '{column}' não foi encontrada no arquivo Excel.")

        df["Preço Base Reais"] = df["Preço Base Original"] * df["Multiplicador Imposto"]

        df["Categoria"] = df["Produtos"].apply(
            lambda x: 'Tecnologia' if 'Tablet' in x or 'Celular' in x or 'Computador' in x else 'Outro'
        )

        def calcular_desconto(tipo_produto):
            if tipo_produto == 'Produto':
                return 0.10  
            elif tipo_produto == 'Serviço':
                return 0.15  
            else:
                return 0.05  

        def calcular_taxa_frete(tipo_produto):
            if tipo_produto == 'Produto':
                return 50  
            elif tipo_produto == 'Serviço':
                return 100  
            else:
                return 20  

        df["Tipo"] = df["Produtos"].apply(
            lambda x: 'Produto' if 'Tablet' in x or 'Celular' in x or 'Computador' in x else 'Serviço'  
        )

        df["Desconto"] = df["Tipo"].apply(calcular_desconto)
        df["Taxa de Frete"] = df["Tipo"].apply(calcular_taxa_frete)

        df["Preço com Desconto"] = df["Preço Base Original"] * (1 - df["Desconto"])
        df["Preço Final"] = df["Preço com Desconto"] + df["Taxa de Frete"]

        df.to_excel("./dados/Produtos(1).xlsx", index=False)

        # openpyxl
        wb = load_workbook("./dados/Produtos(1).xlsx")
        wa = wb.active

        style = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        for i in range(2, wa.max_row + 1):  
            cell = wa[f"F{i}"]  
            if str(cell.value) == 'Tecnologia':  
                cell.fill = style  

        wb.save("./dados/Produtos(1).xlsx")
        print("Dataframe atualizado 🚀")
    
    except Exception as e:
        print(f"Ocorreu um erro: {e}")


schedule.every().monday.at("10:00").do(weekly_data_processing)

while True:
    schedule.run_pending()
    time.sleep(1)