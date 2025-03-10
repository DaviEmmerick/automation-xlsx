from openpyxl import load_workbook
from copy import copy

def create_sheet(district, neighborhoods_file, header_style):
    if district not in neighborhoods_file.sheetnames:
        neighborhoods_file.create_sheet(district)
        new_sheet = neighborhoods_file[district]

        new_sheet["A1"].value = "Data de Nascimento"
        new_sheet["B1"].value = "Pessoa"
        new_sheet["C1"].value = "Bairro"

        new_sheet["A1"]._style = header_style
        new_sheet["B1"]._style = header_style
        new_sheet["C1"]._style = header_style

def transfer_data_from_sheet(source_sheet, destination_sheet, source_row):
    destination_row = destination_sheet.max_row + 1
    for column in range(1, 4):
        source_cell = source_sheet.cell(row=source_row, column=column)
        destination_cell = destination_sheet.cell(row=destination_row, column=column)

        destination_cell.value = source_cell.value
        destination_cell._style = copy(source_cell._style)

def process_neighborhoods_data(file_path: str):
    try:
        neighborhoods_file = load_workbook(file_path)
        base_data_sheet = neighborhoods_file["Base de Dados"]
    except Exception as e:
        print(f"Erro ao carregar o arquivo ou aba 'Base de Dados': {e}")
        return

    last_row = base_data_sheet.max_row
    print(f"Última linha de dados: {last_row}")

    header_style = copy(base_data_sheet["A1"]._style)

    for row in range(2, last_row + 1):
        district = base_data_sheet.cell(row=row, column=3).value
        if not district:
            break

        create_sheet(district, neighborhoods_file, header_style)
        destination_sheet = neighborhoods_file[district]
        transfer_data_from_sheet(base_data_sheet, destination_sheet, row)

    neighborhoods_file.save(file_path)
    print(f"Arquivo {file_path} atualizado com sucesso!")